import os
import sys

path_to_add = os.path.abspath(os.path.join(os.path.dirname(__file__), '../'))
if path_to_add not in sys.path:
    sys.path.insert(0, path_to_add)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from portfolio_tracker import config
from collections import defaultdict

def clean_value(val):
    if isinstance(val, str):
        # Remove $, commas, and whitespace
        val = val.replace('$', '').replace(',', '').strip()
    try:
        return float(val)
    except:
        return 0
    


def generate_portfolio_report(symbol_col, value_col, df):
    # Prepare data for chart (convert values to numbers)
    chart_data = df[[symbol_col, value_col]].copy()
    # Convert to DataFrame for chart
    chart_data = pd.DataFrame(list(categorized_portfolio.items()), columns=['Symbol', 'Value'])

    total = chart_data.iloc[1:, 1].sum()  # Sum of values column
    chart_data['Percentage'] = ((chart_data.iloc[:, 1] / total))
    
    # Export to Excel
    chart_data.to_excel(config.OUTPUT_FILE, index=False, sheet_name='Portfolio')
    
    # Add pie chart
    wb = load_workbook(config.OUTPUT_FILE)
    ws = wb['Portfolio']
    for row in range(2, len(chart_data) + 2):
        cell = ws[f'C{row}']
        cell.number_format = '0.00%'  # Excel percentage format
    # Create pie chart
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(chart_data)+1)  # Column A: Symbols
    data = Reference(ws, min_col=3, min_row=2, max_row=len(chart_data)+1)    # Column C: Percentage

    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Portfolio Distribution"

    # Add percentages to the chart
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = True          # Show the percentage values
    pie.dataLabels.showCatName = False     # Don't show symbol name
    pie.dataLabels.showPercent = False     # Don't calculate percent (we already have it)
    pie.dataLabels.showSerName = False     # Don't show series name (this removes "Series")
    pie.dataLabels.showLeaderLines = False
    pie.dataLabels.showLegendKey = False
    # Position labels outside the pie slices to prevent overlap

    #pie.dataLabels.position = 'bestFit'  # or 'outEnd' to put all labels outside
    pie.dataLabels.position = 'bestFit'
    # Make the chart bigger to fit labels better
    # Make the chart much bigger to fit labels better
    pie.width = 20  # Width in cm
    pie.height = 15  # Height in cm
    # Add chart to sheet
    # Remove chart title (we'll add it as a cell instead)
    pie.title = None

    # Add title as a cell above the chart
    ws['E2'] = 'Portfolio Distribution'
    ws['E2'].font = ws['E2'].font.copy(size=16, bold=True)
    ws.add_chart(pie, "E4")
    
    # Save
    wb.save(config.OUTPUT_FILE)

    print(f"✓ Excel file created: {config.OUTPUT_FILE}")
    print("  - Data is in columns A & B")
    print("  - Pie chart is on the right")



def process_fidelity_data():
    with open(config.INPUT_FILE, 'r') as f:
        lines = f.readlines()

    # Add comma to the first line (header) if it doesn't end with comma
    if lines[0].strip() and not lines[0].strip().endswith(','):
        lines[0] = lines[0].strip() + ',\n'

    # Write the fixed version
    with open(config.TEMP_FILE, 'w') as f:
        f.writelines(lines)

    # Now read the fixed CSV
    df = pd.read_csv(config.TEMP_FILE)

    # Find the Symbol column
    symbol_col = None
    for col in df.columns:
        if 'symbol' in str(col).lower() or 'ticker' in str(col).lower():
            symbol_col = col
            break

    # Find the Value column
    value_col = None
    for col in df.columns:
        if 'value' in str(col).lower() or 'amount' in str(col).lower() or 'total' in str(col).lower():
            value_col = col
            break
    
    return symbol_col, value_col, df


def process_ib_data():
    with open(config.INPUT_FILE_IB, 'r') as f:
        lines = f.readlines()

    # Add comma to the first line (header) if it doesn't end with comma
    if lines[0].strip() and not lines[0].strip().endswith(','):
        lines[0] = lines[0].strip() + ',\n'

    # Write the fixed version
    with open(config.TEMP_FILE, 'w') as f:
        f.writelines(lines)

    # Now read the fixed CSV
    df = pd.read_csv(config.TEMP_FILE)

    # Interactive Brokers uses fixed columns: A=Symbol, B=Value
    symbol_col = df.columns[0]
    value_col = df.columns[1]
    
    return symbol_col, value_col, df




def process_portfolio_data(symbol_col, value_col, df):
    portfolio_data = defaultdict(list)
    if symbol_col and value_col:
        for index, row in df.iterrows():
            symbol = row[symbol_col]
            value = row[value_col]
            portfolio_data[symbol].append(clean_value(value))
        portfolio_data = {key: sum(values) for key, values in portfolio_data.items()}
    else:
        print("Error: Could not find Symbol or Value colulmns")
        print(f"Available columns: {list(df.columns)}")
    
    return portfolio_data

symbol_col, value_col, df = process_fidelity_data()
portfolio_data_fidelity = process_portfolio_data(symbol_col, value_col, df)

symbol_col, value_col, df = process_ib_data()
portfolio_data_ib = process_portfolio_data(symbol_col, value_col, df)

# Concatenate both portfolios
portfolio_data = defaultdict(float)
for portfolio in [portfolio_data_fidelity, portfolio_data_ib]:
    for symbol, value in portfolio.items():
        portfolio_data[symbol] += value

categorized_portfolio = defaultdict(float)
for symbol, value in portfolio_data.items():
    # Initialize and handle NaN values
    if pd.isna(value) or value != value:  # NaN check
        value = 0
    
    category = 'Other'
    # Check all categories in config
    for cat, tickers in config.TICKERS.items():
        if symbol in tickers:
            category = cat
            break
    categorized_portfolio[category] += value
    print("\nPORTFOLIO BY CATEGORY")
    print("=" * 50)
    for category, total in sorted(categorized_portfolio.items()):
        print(f"{category:<20} ${total:>15,.2f}")
    print(f"Total Portfolio: ${sum(categorized_portfolio.values()):>15,.2f}")


generate_portfolio_report(symbol_col, value_col, df)